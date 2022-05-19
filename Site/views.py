from calendar import monthrange
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.utils.decorators import method_decorator
from django.views.generic import *
from openpyxl import load_workbook

from .forms import *
from .functions import *
from .models import *
from django.utils import timezone

now = timezone.now()

errorList = {
    'Error001': "Staff not Found. It appears you do not have an account in the given company",
    'Error002': "Company not Found. It appears this company does not exist",
    'Error003': "PMS not Set. It appears the company has not set PMS for the year or is deactivated",
    'Error004': "Permissions Error. It Appears you do not have permissions for this page",
}


def excel_to_db():
    file = "kpi_records.xlsx"
    sheet = 'cfaokenya'

    wb = load_workbook(file, data_only=True)
    sh = wb[sheet]
    record = 2

    def check_value(x):
        x = str(x).strip()
        if x == "None" or x == "NULL" or x is None or x == "":
            return None
        else:
            return x

    def check_number_value(x):
        x = str(x).strip()
        if x == "None" or x == "NULL" or x is None or x == "":
            return None
        else:
            return int(x)
        
    def check_float_value(x):
        x = str(x).strip()
        if x == "None" or x == "NULL" or x is None or x == "":
            return None
        else:
            return float(x)

    def check_if_true(x):
        if int(x) == 1:
            return True
        else:
            return False

    def approval_date(x):
        x = str(x).strip()
        if x == "None" or x == "NULL" or x is None or x == "":
            return None
        else:
            return datetime.datetime.strptime(x, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.get_current_timezone())

    tz = timezone.get_current_timezone()

    records = 2291
    while record <= records:
        kpi = KPI()
        kpi.kpi_id = sh["A" + str(record)].value
        kpi.kpi_title = sh["B" + str(record)].value
        kpi.kpi_details = check_value(sh["C" + str(record)].value)
        kpi.kpi_criteria = None
        kpi.kpi_target = 0
        kpi.kpi_weight = 20
        kpi.kpi_units = check_value(sh["G" + str(record)].value)
        kpi.kpi_function = check_value(sh["H" + str(record)].value)

        kpi.kpi_april_target = sh["I" + str(record)].value
        kpi.kpi_may_target = sh["J" + str(record)].value
        kpi.kpi_june_target = sh["K" + str(record)].value
        kpi.kpi_july_target = sh["L" + str(record)].value
        kpi.kpi_august_target = sh["M" + str(record)].value
        kpi.kpi_september_target = sh["N" + str(record)].value
        kpi.kpi_october_target = sh["O" + str(record)].value
        kpi.kpi_november_target = sh["P" + str(record)].value
        kpi.kpi_december_target = sh["Q" + str(record)].value
        kpi.kpi_january_target = sh["R" + str(record)].value
        kpi.kpi_february_target = sh["S" + str(record)].value
        kpi.kpi_march_target = sh["T" + str(record)].value

        kpi.kpi_april_score = None
        kpi.kpi_may_score = None
        kpi.kpi_june_score = None
        kpi.kpi_july_score = None
        kpi.kpi_august_score = check_float_value(sh["Y" + str(record)].value)
        kpi.kpi_september_score = check_float_value(sh["Z" + str(record)].value)
        kpi.kpi_october_score = check_float_value(sh["AA" + str(record)].value)
        kpi.kpi_november_score = check_float_value(sh["AB" + str(record)].value)
        kpi.kpi_december_score = check_float_value(sh["AC" + str(record)].value)
        kpi.kpi_january_score = check_float_value(sh["AD" + str(record)].value)
        kpi.kpi_february_score = check_float_value(sh["AE" + str(record)].value)
        kpi.kpi_march_score = check_float_value(sh["AF" + str(record)].value)

        kpi.kpi_bsc_s_target = None
        kpi.kpi_bsc_a_target = None
        kpi.kpi_bsc_b_target = None
        kpi.kpi_bsc_c_target = None
        kpi.kpi_bsc_d_target = None

        kpi.kpi_april_score_approve = check_if_true(sh["AL" + str(record)].value)
        kpi.kpi_may_score_approve = check_if_true(sh["AM" + str(record)].value)
        kpi.kpi_june_score_approve = check_if_true(sh["AN" + str(record)].value)
        kpi.kpi_july_score_approve = check_if_true(sh["AO" + str(record)].value)
        kpi.kpi_august_score_approve = check_if_true(sh["AP" + str(record)].value)
        kpi.kpi_september_score_approve = check_if_true(sh["AQ" + str(record)].value)
        kpi.kpi_october_score_approve = check_if_true(sh["AR" + str(record)].value)
        kpi.kpi_november_score_approve = check_if_true(sh["AS" + str(record)].value)
        kpi.kpi_december_score_approve = check_if_true(sh["AT" + str(record)].value)
        kpi.kpi_january_score_approve = check_if_true(sh["AU" + str(record)].value)
        kpi.kpi_february_score_approve = check_if_true(sh["AV" + str(record)].value)
        kpi.kpi_march_score_approve = check_if_true(sh["AW" + str(record)].value)

        kpi.kpi_submit_date = sh["AX" + str(record)].value
        kpi.kpi_status = sh["AY" + str(record)].value
        kpi.kpi_type = "Average"
        kpi.kpi_all_results_approve = False
        kpi.kpi_bsc_pillar = None
        kpi.kpi_pms_id = "44c5c6383e524f1d9e5646534f9faab5"
        kpi.kpi_staff_id = sh["BD" + str(record)].value
        kpi.save()

        # print("Done - " + str(get_object_or_404(Staff, staff_id=int(sh["BD" + str(record)].value))) + " " + str(record))
        record += 1


def update_to_db():
    file = "tken_update_kpi.xlsx"
    sheet = 'cfaokenya'

    wb = load_workbook(file, data_only=True)
    sh = wb[sheet]
    record = 2

    def check_float_value(x):
        x = str(x).strip()
        if x == "None" or x == "NULL" or x is None or x == "":
            return None
        else:
            return float(x)

    records = 142
    while record <= records:
        kpi = KPI.objects.get(kpi_id=sh["A" + str(record)].value, kpi_staff_id=sh["N" + str(record)].value)
        kpi.kpi_april_target = check_float_value(sh["B" + str(record)].value)
        kpi.kpi_may_target = check_float_value(sh["C" + str(record)].value)
        kpi.kpi_june_target = check_float_value(sh["D" + str(record)].value)
        kpi.kpi_july_target = check_float_value(sh["E" + str(record)].value)
        kpi.kpi_august_target = check_float_value(sh["F" + str(record)].value)
        kpi.kpi_september_target = check_float_value(sh["G" + str(record)].value)
        kpi.kpi_october_target = check_float_value(sh["H" + str(record)].value)
        kpi.kpi_november_target = check_float_value(sh["I" + str(record)].value)
        kpi.kpi_december_target = check_float_value(sh["J" + str(record)].value)
        kpi.kpi_january_target = check_float_value(sh["K" + str(record)].value)
        kpi.kpi_february_target = check_float_value(sh["L" + str(record)].value)
        kpi.kpi_march_target = check_float_value(sh["M" + str(record)].value)
        kpi.save()

        # print("Done - " + str(get_object_or_404(Staff, staff_id=int(sh["N" + str(record)].value))) + " " + str(record))
        record += 1


@method_decorator(login_required, name='dispatch')
@method_decorator(password_change_decorator, name='dispatch')
class Dashboard(TemplateView):
    def get_context_data(self, **kwargs):
        context = super(Dashboard, self).get_context_data()
        context['company_id'] = self.kwargs['company_id']

        context['error_code'] = checks(self.kwargs['company_id'], self.request.user)
        global_context(self.kwargs['company_id'], self.request.user, context)

        kpi_type = KPIType.objects.filter(type_pms=context['pms'],
                                          type_category=get_staff_account(get_company(self.kwargs['company_id']),
                                                                          self.request.user).staff_category)
        if kpi_type:
            context['kpi_type'] = kpi_type.first().type_kpi
        else:
            context['kpi_type'] = "Annual Target"

        kpi_results = []
        if context['pms']:
            context['kpi_overall_results'] = calculate_overall_kpi_score(context['staff'], context['pms'])
            context['checkin_overall_results'] = calculate_overall_check_in_score(context['staff'], context['pms'])
            context['checkin'] = get_user_checkin(context['staff'], context['pms'])
            context['assessment_overall_score'] = calculate_overall_assessment_score(context['staff'], context['pms'])
            context['overall_score'] = calculate_overall_score(context['staff'], context['pms'])
            context['matrix_applied'] = display_matrix(context['staff'], context['pms'])
            context['company_score'] = get_company_score(context['staff'], context['pms'])
            context['kpi_months_used'] = display_months_used(context['staff'], context['pms'])
            context['assessment_used'] = Assessment.objects.filter(assessment_pms=context['pms'])
            context['checkin_used'] = display_checkin_scoring_used(context['staff'], context['pms'])
            context['bu_score'] = get_bu_score(context['staff'], context['pms'])
            context['the_bu'] = get_bu(context['staff'])
            context['the_company'] = get_company_used(context['staff'])

            for kpi in KPI.objects.filter(kpi_staff=context['staff'], kpi_pms=context['pms']):
                kpi_results.append([kpi, calculate_kpi_score2(kpi, context['kpi_type'])])
        context['kpi_results'] = kpi_results

        return context


# Error
@method_decorator(login_required, name='dispatch')
class ErrorPage(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(ErrorPage, self).get_context_data()
        context['error_code'] = self.kwargs['error_code']
        context['error_detail'] = errorList[self.kwargs['error_code']]
        return context


@method_decorator(login_required, name='dispatch')
class MyKPI(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(MyKPI, self).get_context_data()
        context['company_id'] = self.kwargs['company_id']
        context['error_code'] = checks(self.kwargs['company_id'], self.request.user)
        global_context(self.kwargs['company_id'], self.request.user, context)
        context['date_check'] = datetime.datetime.now()
        context = context | kpi_list(context['staff'], context['pms'])

        return context


@method_decorator(login_required, name='dispatch')
class MyKPICreate(CreateView):
    form_class = KPIForm

    def get_context_data(self, **kwargs):
        context = super(MyKPICreate, self).get_context_data()
        context['company_id'] = self.kwargs['company_id']
        context['error_code'] = checks(self.kwargs['company_id'], self.request.user)
        global_context(self.kwargs['company_id'], self.request.user, context)

        if context['staff'] and context['company'] and context['pms']:
            context = context | kpi_submission_checks(context['staff'], context['pms'])
            context = context | kpi_list(context['staff'], context['pms'])
            context['number_kpis'] = kpi_number_check(context['staff'], context['pms'])
            context['sum_weight'] = kpi_weight_check(context['staff'], context['pms'])
            context['submission_data'] = get_user_submission_data(context['staff'], context['pms'])

            kpi_type = KPIType.objects.filter(type_pms=context['pms'],
                                              type_category=get_staff_account(get_company(self.kwargs['company_id']),
                                                                              self.request.user).staff_category)
            if kpi_type:
                context['kpi_type'] = kpi_type.first().type_kpi
            else:
                context['kpi_type'] = "Annual Target"

        return context

    def get_success_url(self):
        return '{}'.format(reverse('Site:My_KPI_Create', kwargs={'company_id': self.kwargs['company_id']}))

    def get_initial(self):
        initial = super(MyKPICreate, self).get_initial()
        initial['kpi_staff'] = get_staff_account(get_company(self.kwargs['company_id']), self.request.user)
        initial['kpi_pms'] = get_active_pms(get_company(self.kwargs['company_id']))
        initial['kpi_submit_date'] = datetime.datetime.now()

        return initial

    def form_valid(self, form):
        super(MyKPICreate, self).form_valid(form)
        staff = get_staff_account(get_company(self.kwargs['company_id']), self.request.user)

        notification_log("KPI", "None", staff.staff_person.get_full_name(), staff.staff_person.email,
                         "My KPI: Submitted",
                         "It appears you have submitted a KPI ")
        if get_staff_level(staff) is not None:
            level = get_staff_level(staff)
            notification_log("KPI", "None", level.level_head.staff_person.get_full_name(),
                             level.level_head.staff_person.email,
                             "Staff KPI: Submitted",
                             "It appears " + staff.staff_person.get_full_name() +
                             " has submitted a KPI for your review")
        return HttpResponseRedirect(reverse('Site:My_KPI_Create', kwargs={'company_id': self.kwargs['company_id']}))


@method_decorator(login_required, name='dispatch')
class MyKPIView(DetailView):
    model = KPI

    def get_context_data(self, **kwargs):
        context = super(MyKPIView, self).get_context_data()
        context['company_id'] = self.kwargs['company_id']
        context['error_code'] = checks(self.kwargs['company_id'], self.request.user)
        global_context(self.kwargs['company_id'], self.request.user, context)

        kpi_type = KPIType.objects.filter(type_pms=context['pms'],
                                          type_category=get_staff_account(get_company(self.kwargs['company_id']),
                                                                          self.request.user).staff_category)
        if kpi_type:
            context['kpi_type'] = kpi_type.first().type_kpi
        else:
            context['kpi_type'] = "Annual Target"

        return context


@method_decorator(login_required, name='dispatch')
class MyKPIEdit(UpdateView):
    model = KPI
    form_class = KPIForm

    def get_context_data(self, **kwargs):
        context = super(MyKPIEdit, self).get_context_data()
        context['company_id'] = self.kwargs['company_id']
        context['error_code'] = checks(self.kwargs['company_id'], self.request.user)
        global_context(self.kwargs['company_id'], self.request.user, context)

        kpi_type = KPIType.objects.filter(type_pms=context['pms'],
                                          type_category=get_staff_account(get_company(self.kwargs['company_id']),
                                                                          self.request.user).staff_category)
        if kpi_type:
            context['kpi_type'] = kpi_type.first().type_kpi
        else:
            context['kpi_type'] = "Annual Target"

        return context

    def get_initial(self):
        initial = super(MyKPIEdit, self).get_initial()
        initial['kpi_staff'] = get_staff_account(get_company(self.kwargs['company_id']), self.request.user)
        initial['kpi_pms'] = get_active_pms(get_company(self.kwargs['company_id']))
        initial['kpi_status'] = 'Submitted'

        return initial

    def get_success_url(self):
        return '{}'.format(reverse('Site:My_KPI_View', kwargs={'company_id': self.kwargs['company_id'],
                                                               'pk': self.kwargs['pk']}))

    def form_valid(self, form):
        super(MyKPIEdit, self).form_valid(form)
        staff = get_staff_account(get_company(self.kwargs['company_id']), self.request.user)
        kpi = get_object_or_404(KPI, kpi_id=self.kwargs['pk'])
        notification_log("KPI", self.kwargs['pk'], staff.staff_person.get_full_name(), staff.staff_person.email,
                         "My KPI: " + kpi.kpi_title,
                         "It appears you have modified or populated results for KPI " + kpi.kpi_title)
        if get_staff_level(staff) is not None:
            level = get_staff_level(staff)
            notification_log("KPI", self.kwargs['pk'], level.level_head.staff_person.get_full_name(),
                             level.level_head.staff_person.email,
                             "Staff KPI: " + kpi.kpi_title,
                             "It appears " + staff.staff_person.get_full_name() +
                             " has modified or populated results for KPI " + kpi.kpi_title +
                             " for your review")
        return HttpResponseRedirect(reverse('Site:My_KPI_View', kwargs={'company_id': self.kwargs['company_id'],
                                                                        'pk': self.kwargs['pk']}))


@method_decorator(login_required, name='dispatch')
class MyKPIDelete(DeleteView):
    model = KPI

    def get_context_data(self, **kwargs):
        context = super(MyKPIDelete, self).get_context_data()
        context['company_id'] = self.kwargs['company_id']
        context['error_code'] = checks(self.kwargs['company_id'], self.request.user)
        global_context(self.kwargs['company_id'], self.request.user, context)
        return context

    def get_success_url(self):
        return '{}'.format(reverse('Site:My_KPI', kwargs={'company_id': self.kwargs['company_id']}))


@method_decorator(login_required, name='dispatch')
class MyKPIResults(UpdateView):
    model = KPI
    form_class = KPIForm

    def get_context_data(self, **kwargs):
        kpi = get_object_or_404(KPI, kpi_id=self.kwargs['pk'])
        context = super(MyKPIResults, self).get_context_data()
        context['company_id'] = self.kwargs['company_id']
        context['error_code'] = checks(self.kwargs['company_id'], self.request.user)
        global_context(self.kwargs['company_id'], self.request.user, context)
        context = context | kpi_list(context['staff'], context['pms'])

        kpi_type = KPIType.objects.filter(type_pms=context['pms'],
                                          type_category=get_staff_account(get_company(self.kwargs['company_id']),
                                                                          self.request.user).staff_category)
        if kpi_type:
            context['kpi_type'] = kpi_type.first().type_kpi
        else:
            context['kpi_type'] = "Annual Target"

        months = {}
        reveal = {}
        if context['pms']:
            year = context['calendar_dict']
            context = context | kpi_list(context['staff'], context['pms'])
            submission = get_user_submission_data(context['staff'], context['pms'])

            if submission:
                months['April'] = submission.submission_april_results
                months['May'] = submission.submission_may_results
                months['June'] = submission.submission_june_results
                months['July'] = submission.submission_july_results
                months['August'] = submission.submission_august_results
                months['September'] = submission.submission_september_results
                months['October'] = submission.submission_october_results
                months['November'] = submission.submission_november_results
                months['December'] = submission.submission_december_results
                months['January'] = submission.submission_january_results
                months['February'] = submission.submission_february_results
                months['March'] = submission.submission_march_results

                context['april_override'] = submission.submission_april_results_override
                context['may_override'] = submission.submission_may_results_override
                context['june_override'] = submission.submission_june_results_override
                context['july_override'] = submission.submission_july_results_override
                context['august_override'] = submission.submission_august_results_override
                context['september_override'] = submission.submission_september_results_override
                context['october_override'] = submission.submission_october_results_override
                context['november_override'] = submission.submission_november_results_override
                context['december_override'] = submission.submission_december_results_override
                context['january_override'] = submission.submission_january_results_override
                context['february_override'] = submission.submission_february_results_override
                context['march_override'] = submission.submission_march_results_override

            else:
                months['April'] = months['May'] = months['June'] = months['July'] = months['August'] = \
                    months['September'] = months['October'] = months['November'] = months['December'] = \
                    months['January'] = months['February'] = months['March'] = 15

                context['april_override'] = False
                context['may_override'] = False
                context['june_override'] = False
                context['july_override'] = False
                context['august_override'] = False
                context['september_override'] = False
                context['october_override'] = False
                context['november_override'] = False
                context['december_override'] = False
                context['january_override'] = False
                context['february_override'] = False
                context['march_override'] = False

            # check if the field result field should show

            # 1. Check if result is approved
            if kpi.kpi_all_results_approve is True:
                reveal['April'] = reveal['May'] = reveal['June'] = reveal['July'] = reveal['August'] = \
                    reveal['September'] = reveal['October'] = reveal['November'] = reveal['December'] = \
                    reveal['January'] = reveal['February'] = reveal['March'] = False
            else:
                today_date = datetime.date.today()
                reveal['April'] = reveal['May'] = reveal['June'] = reveal['July'] = reveal['August'] = \
                    reveal['September'] = reveal['October'] = reveal['November'] = reveal['December'] = \
                    reveal['January'] = reveal['February'] = reveal['March'] = False

                # Check if the result is approved or within the time

                # April Check
                if kpi.kpi_april_score_approve is True:
                    reveal['April'] = False
                else:
                    if context['april_override'] is True:
                        reveal['April'] = True
                    else:
                        april_end_month = datetime.date(year=year['April'], month=4,
                                                        day=monthrange(year['April'], 4)[1])
                        april_deadline = april_end_month + datetime.timedelta(days=months['April'])

                        if april_end_month <= today_date <= april_deadline:
                            reveal['April'] = True

                # May Check
                if kpi.kpi_may_score_approve is True:
                    reveal['May'] = False
                else:
                    if context['may_override'] == True:
                        reveal['May'] = True
                    else:
                        may_end_month = datetime.date(year=year['May'], month=5, day=monthrange(year['May'], 5)[1])
                        may_deadline = may_end_month + datetime.timedelta(days=months['May'])

                        if may_end_month <= today_date <= may_deadline:
                            reveal['May'] = True

                # June Check
                if kpi.kpi_june_score_approve is True:
                    reveal['June'] = False
                else:
                    if context['june_override'] == True:
                        reveal['June'] = True
                    else:
                        june_end_month = datetime.date(year=year['June'], month=6, day=monthrange(year['June'], 6)[1])
                        june_deadline = june_end_month + datetime.timedelta(days=months['June'])

                        if june_end_month <= today_date <= june_deadline:
                            reveal['June'] = True

                # July Check
                if kpi.kpi_july_score_approve is True:
                    reveal['July'] = False
                else:
                    if context['july_override'] == True:
                        reveal['June'] = True
                    else:
                        july_end_month = datetime.date(year=year['July'], month=7, day=monthrange(year['July'], 7)[1])
                        july_deadline = july_end_month + datetime.timedelta(days=months['July'])

                        if july_end_month <= today_date <= july_deadline:
                            reveal['July'] = True

                # August Check
                if kpi.kpi_august_score_approve is True:
                    reveal['August'] = False
                else:
                    if context['august_override'] == True:
                        reveal['August'] = True
                    else:
                        august_end_month = datetime.date(year=year['August'], month=8, day=monthrange(year['August'], 8)[1])
                        august_deadline = august_end_month + datetime.timedelta(days=months['August'])

                        if august_end_month <= today_date <= august_deadline:
                            reveal['August'] = True

                # September Check
                if kpi.kpi_september_score_approve is True:
                    reveal['September'] = False
                else:
                    if context['september_override'] == True:
                        reveal['September'] = True
                    else:
                        september_end_month = datetime.date(year=year['September'], month=9,
                                                            day=monthrange(year['September'], 9)[1])
                        september_deadline = september_end_month + datetime.timedelta(days=months['September'])

                        if september_end_month <= today_date <= september_deadline:
                            reveal['September'] = True

                # October Check
                if kpi.kpi_october_score_approve is True:
                    reveal['October'] = False
                else:
                    if context['october_override'] == True:
                        reveal['October'] = True
                    else:
                        october_end_month = datetime.date(year=year['October'], month=10,
                                                          day=monthrange(year['October'], 10)[1])
                        october_deadline = october_end_month + datetime.timedelta(days=months['October'])

                        if october_end_month <= today_date <= october_deadline:
                            reveal['October'] = True

                # November Check
                if kpi.kpi_november_score_approve is True:
                    reveal['November'] = False
                else:
                    if context['november_override'] == True:
                        reveal['November'] = True
                    else:
                        november_end_month = datetime.date(year=year['November'], month=11,
                                                           day=monthrange(year['November'], 11)[1])
                        november_deadline = november_end_month + datetime.timedelta(days=months['November'])

                        if november_end_month <= today_date <= november_deadline:
                            reveal['November'] = True

                # December Check
                if kpi.kpi_december_score_approve is True:
                    reveal['December'] = False
                else:
                    if context['december_override'] == True:
                        reveal['December'] = True
                    else:
                        december_end_month = datetime.date(year=year['December'], month=12,
                                                           day=monthrange(year['December'], 12)[1])
                        december_deadline = december_end_month + datetime.timedelta(days=months['December'])

                        if december_end_month <= today_date <= december_deadline:
                            reveal['December'] = True

                # January Check
                if kpi.kpi_january_score_approve is True:
                    reveal['January'] = False
                else:
                    if context['january_override'] == True:
                        reveal['January'] = True
                    else:
                        january_end_month = datetime.date(year=year['January'], month=1,
                                                          day=monthrange(year['January'], 1)[1])
                        january_deadline = january_end_month + datetime.timedelta(days=months['January'])

                        if january_end_month <= today_date <= january_deadline:
                            reveal['January'] = True

                # February Check
                if kpi.kpi_february_score_approve is True:
                    reveal['February'] = False
                else:
                    if context['february_override'] == True:
                        reveal['February'] = True
                    else:
                        february_end_month = datetime.date(year=year['February'], month=2,
                                                           day=monthrange(year['February'], 2)[1])
                        february_deadline = february_end_month + datetime.timedelta(days=months['February'])

                        if february_end_month <= today_date <= february_deadline:
                            reveal['February'] = True

                # March Check
                if kpi.kpi_march_score_approve is True:
                    reveal['March'] = False
                else:
                    if context['march_override'] == True:
                        reveal['March'] = True
                    else:
                        march_end_month = datetime.date(year=year['March'], month=3, day=monthrange(year['March'], 3)[1])
                        march_deadline = march_end_month + datetime.timedelta(days=months['March'])

                        if march_end_month <= today_date <= march_deadline:
                            reveal['March'] = True

        context['reveal'] = reveal
        return context

    def get_initial(self):
        initial = super(MyKPIResults, self).get_initial()
        initial['kpi_staff'] = get_staff_account(get_company(self.kwargs['company_id']), self.request.user)
        initial['kpi_pms'] = get_active_pms(get_company(self.kwargs['company_id']))

        return initial

    def get_success_url(self):
        return '{}'.format(reverse('Site:My_KPI_Results', kwargs={'company_id': self.kwargs['company_id'],
                                                                  'pk': self.kwargs['pk']}))

    def form_valid(self, form):
        super(MyKPIResults, self).form_valid(form)
        staff = get_staff_account(get_company(self.kwargs['company_id']), self.request.user)
        kpi = get_object_or_404(KPI, kpi_id=self.kwargs['pk'])
        notification_log("KPI", self.kwargs['pk'], staff.staff_person.get_full_name(), staff.staff_person.email,
                         "My KPI: " + kpi.kpi_title,
                         "It appears you have modified or populated results for KPI " + kpi.kpi_title )
        if get_staff_level(staff) is not None:
            level = get_staff_level(staff)
            notification_log("KPI", self.kwargs['pk'], level.level_head.staff_person.get_full_name(),
                             level.level_head.staff_person.email,
                             "Staff KPI: " + kpi.kpi_title,
                             "It appears " + staff.staff_person.get_full_name() +
                             " has modified or populated results for KPI " + kpi.kpi_title +
                             " for your review")
        return HttpResponseRedirect(reverse('Site:My_KPI_Results', kwargs={'company_id': self.kwargs['company_id'],
                                                                           'pk': self.kwargs['pk']}))


@method_decorator(login_required, name='dispatch')
class MyKPIResultsList(TemplateView):
    def get_context_data(self, **kwargs):
        context = super(MyKPIResultsList, self).get_context_data()

        context['company_id'] = self.kwargs['company_id']
        context['error_code'] = checks(self.kwargs['company_id'], self.request.user)
        global_context(self.kwargs['company_id'], self.request.user, context)
        context = context | kpi_list(context['staff'], context['pms'])

        kpi_type = KPIType.objects.filter(type_pms=context['pms'],
                                          type_category=get_staff_account(get_company(self.kwargs['company_id']),
                                                                          self.request.user).staff_category)
        if kpi_type:
            context['kpi_type'] = kpi_type.first().type_kpi
        else:
            context['kpi_type'] = "Annual Target"
        return context


@method_decorator(login_required, name='dispatch')
class KPILevelDown(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(KPILevelDown, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        context['level_head'] = Level.objects.filter(level_head=context['staff'],
                                                     level_category__category_company__company_id=context['company_id'])
        levels_down = []
        for level in context['level_head']:
            all_levels_down(level, levels_down)
        context['levels_down'] = levels_down

        return context


@method_decorator(login_required, name='dispatch')
class KPILevelDownDetail(DetailView):
    model = Level

    def get_context_data(self, **kwargs):
        context = super(KPILevelDownDetail, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)

        level = get_object_or_404(Level, level_id=self.kwargs['pk'])
        context['staff_members'] = LevelMembership.objects.filter(membership_level=level)
        context['staff_members_active'] = LevelMembership.objects.filter(membership_is_active=True,
                                                                         membership_level=level)
        context['staff_members_inactive'] = LevelMembership.objects.filter(membership_is_active=False,
                                                                           membership_level=level)

        members = []
        for member in LevelMembership.objects.filter(membership_level=level):
            kpi = kpi_list(member.membership_staff, context['pms'])

            submitted_kpi = kpi['submitted_kpi'].count()
            approved_kpi = kpi['approved_kpi'].count()
            pending_kpi = kpi['pending_kpi'].count()

            members.append([member.membership_staff, submitted_kpi, pending_kpi, approved_kpi,
                            calculate_overall_kpi_score(member.membership_staff,
                                                        context['pms'])])
        context['members'] = members

        return context


@method_decorator(login_required, name='dispatch')
class KPILevelDownDetailStaff(DetailView):
    model = Staff

    def get_context_data(self, **kwargs):
        context = super(KPILevelDownDetailStaff, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)

        context['level'] = get_object_or_404(Level, level_id=self.kwargs['lev_id'])
        context['select_staff'] = get_object_or_404(Staff, staff_id=self.kwargs['pk'])

        kpi_results = []
        kpi_type = KPIType.objects.filter(type_pms=context['pms'],
                                          type_category=get_staff_account(get_company(self.kwargs['company_id']),
                                                                          context['select_staff'].staff_person).staff_category)
        if kpi_type:
            context['kpi_type'] = kpi_type.first().type_kpi
        else:
            context['kpi_type'] = "Annual Target"

        for kpi in KPI.objects.filter(kpi_staff=context['select_staff'], kpi_pms=context['pms']):
            kpi_results.append([kpi, calculate_kpi_score(kpi, context['kpi_type'])])
        context['kpi_results'] = kpi_results
        context['kpi_overall_results'] = calculate_overall_kpi_score(context['select_staff'], context['pms'])
        context['kpis'] = KPI.objects.filter(kpi_staff=context['select_staff'], kpi_pms=context['pms'])

        return context


def staff_kpi_results(request, company_id, lev_id, staff_id, pk):
    if request.method == "POST":

        kpi_id = request.POST.get("kpi_id")

        def get_score_value(x):
            if str(x).strip() == "":
                return None
            else:
                return float(x)

        april_score = get_score_value(request.POST.get("april_score"))
        may_score = get_score_value(request.POST.get("may_score"))
        june_score = get_score_value(request.POST.get("june_score"))
        july_score = get_score_value(request.POST.get("july_score"))
        august_score = get_score_value(request.POST.get("august_score"))
        september_score = get_score_value(request.POST.get("september_score"))
        october_score = get_score_value(request.POST.get("october_score"))
        november_score = get_score_value(request.POST.get("november_score"))
        december_score = get_score_value(request.POST.get("december_score"))
        january_score = get_score_value(request.POST.get("january_score"))
        february_score = get_score_value(request.POST.get("february_score"))
        march_score = get_score_value(request.POST.get("march_score"))

        def get_value(x):
            if x == "on":
                return True
            else:
                return False

        april_score_approve = get_value(request.POST.get("april_score_approve"))
        may_score_approve = get_value(request.POST.get("may_score_approve"))
        june_score_approve = get_value(request.POST.get("june_score_approve"))
        july_score_approve = get_value(request.POST.get("july_score_approve"))
        august_score_approve = get_value(request.POST.get("august_score_approve"))
        september_score_approve = get_value(request.POST.get("september_score_approve"))
        october_score_approve = get_value(request.POST.get("october_score_approve"))
        november_score_approve = get_value(request.POST.get("november_score_approve"))
        december_score_approve = get_value(request.POST.get("december_score_approve"))
        january_score_approve = get_value(request.POST.get("january_score_approve"))
        february_score_approve = get_value(request.POST.get("february_score_approve"))
        march_score_approve = get_value(request.POST.get("march_score_approve"))

        if int(pk) == int(kpi_id):
            kpi = get_object_or_404(KPI, kpi_id=pk)

            kpi.kpi_april_score = april_score
            kpi.kpi_april_score_approve = april_score_approve

            kpi.kpi_may_score = may_score
            kpi.kpi_may_score_approve = may_score_approve

            kpi.kpi_june_score = june_score
            kpi.kpi_june_score_approve = june_score_approve

            kpi.kpi_july_score = july_score
            kpi.kpi_july_score_approve = july_score_approve

            kpi.kpi_august_score = august_score
            kpi.kpi_august_score_approve = august_score_approve

            kpi.kpi_september_score = september_score
            kpi.kpi_september_score_approve = september_score_approve

            kpi.kpi_october_score = october_score
            kpi.kpi_october_score_approve = october_score_approve

            kpi.kpi_november_score = november_score
            kpi.kpi_november_score_approve = november_score_approve

            kpi.kpi_december_score = december_score
            kpi.kpi_december_score_approve = december_score_approve

            kpi.kpi_january_score = january_score
            kpi.kpi_january_score_approve = january_score_approve

            kpi.kpi_february_score = february_score
            kpi.kpi_february_score_approve = february_score_approve

            kpi.kpi_march_score = march_score
            kpi.kpi_march_score_approve = march_score_approve

            kpi.save()

            team_leader = get_staff_account(get_company(company_id), request.user)
            staff_member = get_object_or_404(Staff, staff_id=staff_id)
            notification_log("KPI", pk, team_leader.staff_person.get_full_name(), team_leader.staff_person.email,
                             "Staff KPI: " + kpi.kpi_title,
                             "It appears you have modified or approved results for KPI " + kpi.kpi_title +
                             " for user " + staff_member.staff_person.get_full_name())

            notification_log("KPI", pk, staff_member.staff_person.get_full_name(),
                             staff_member.staff_person.email,
                             "My KPI: " + kpi.kpi_title,
                             "It appears " + kpi.kpi_title +
                             " results have been modified or approved for KPI " + kpi.kpi_title + " by " +
                             team_leader.staff_person.get_full_name())

    return HttpResponseRedirect(reverse('Site:KPI_LevelDownDetailStaff',
                                        kwargs={'company_id': company_id, 'lev_id': lev_id, 'pk': staff_id}))


@method_decorator(login_required, name='dispatch')
class KPICategory(DetailView):
    model = LevelCategory

    def get_context_data(self, **kwargs):
        context = super(KPICategory, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        context['level_category'] = get_object_or_404(LevelCategory, category_id=self.kwargs['pk'])
        categories = []
        levels_up = []
        levels_in_category = []

        staff_level = get_staff_level(context['staff'])
        if staff_level:
            if staff_level.level_category.category_kpi_view is True and not staff_level.level_head == context[
                'staff'] and staff_level.level_category == context['level_category']:
                levels_in_category.append(staff_level)
            all_levels_up(staff_level, levels_up)

        levels = Level.objects.filter(level_category=context['level_category'])
        for level in levels:
            if level in levels_up:
                levels_in_category.append(level)

        context['level_in_category'] = levels_in_category

        return context


@method_decorator(login_required, name='dispatch')
class KPICategoryLevel(DetailView):
    model = Level

    def get_context_data(self, **kwargs):
        context = super(KPICategoryLevel, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        context['level_category'] = get_object_or_404(LevelCategory, category_id=self.kwargs['cat_id'])

        level = get_object_or_404(Level, level_id=self.kwargs['pk'])

        kpi_results = []
        kpi_type = KPIType.objects.filter(type_pms=context['pms'],
                                          type_category=get_staff_account(get_company(self.kwargs['company_id']),
                                                                          self.request.user).staff_category)
        if kpi_type:
            context['kpi_type'] = kpi_type.first().type_kpi
        else:
            context['kpi_type'] = "Annual Target"

        for kpi in KPI.objects.filter(kpi_staff=level.level_head, kpi_pms=context['pms']):
            kpi_results.append([kpi, calculate_kpi_score2(kpi, context['kpi_type'])])
        context['kpi_results'] = kpi_results
        context['kpi_overall_results'] = calculate_overall_kpi_score(level.level_head, context['pms'])
        context['kpis'] = KPI.objects.filter(kpi_staff=level.level_head, kpi_pms=context['pms'])
        return context


@method_decorator(login_required, name='dispatch')
class MyCheckIn(TemplateView):
    model = CheckIn
    context_object_name = 'CheckIn'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(MyCheckIn, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)

        ci = CheckIn.objects.filter(check_in_Staff=context['staff'])
        context['approved_ci'] = ci.filter(check_in_status='Approved')
        context['pending_ci'] = ci.filter(check_in_status='Pending')
        context['rejected_ci'] = ci.filter(check_in_status='Rejected')
        context['CheckIn'] = CheckIn.objects.filter(check_in_Staff=context['staff'])

        return context


@method_decorator(login_required, name='dispatch')
class MyCheckInCreate(CreateView):
    form_class = CheckInForm

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(MyCheckInCreate, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)

        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month=datetime.date.today().strftime('%B')):
            context['done'] = True
        else:
            context['done'] = False
        context['CheckIn'] = CheckIn.objects.filter(check_in_Staff=context['staff'])
        context['today'] = datetime.date.today()
        return context

    def get_success_url(self):
        return '{}'.format(reverse('Site:My_CheckIn', kwargs={'company_id': self.kwargs['company_id']}))

    def get_initial(self):
        initial = super(MyCheckInCreate, self).get_initial()
        context = {}
        global_context(self.kwargs['company_id'], self.request.user, context)

        initial['check_in_pms'] = context['pms']
        initial['check_in_Staff'] = context['staff']
        initial['check_in_submit_date'] = datetime.datetime.now()
        initial['check_in_month'] = datetime.date.today().strftime('%B')
        return initial

    def form_valid(self, form):
        super(MyCheckInCreate, self).form_valid(form)
        staff = get_staff_account(get_company(self.kwargs['company_id']), self.request.user)

        notification_log("CheckIn", "None", staff.staff_person.get_full_name(), staff.staff_person.email,
                         "My CheckIn Submit", "It appears you have submitted a Check-In")
        if get_staff_level(staff) is not None:
            level = get_staff_level(staff)
            notification_log("CheckIn", "None", level.level_head.staff_person.get_full_name(),
                             level.level_head.staff_person.email, "Staff CheckIn Submitted",
                             "It appears " + staff.staff_person.get_full_name() +
                             " has submitted a checkin for your review")

        return HttpResponseRedirect(reverse('Site:My_CheckIn_Create', kwargs={'company_id': self.kwargs['company_id']}))


@method_decorator(login_required, name='dispatch')
class MyCheckInPrevious(TemplateView):

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(MyCheckInPrevious, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        
        context['April'] = False
        context['May'] = False
        context['June'] = False
        context['July'] = False
        context["August"] = False
        context['September'] = False
        context['October'] = False
        context['November'] = False
        context['December'] = False
        context['January'] = False
        context['February'] = False
        context['March'] = False

        submissions = SubmissionCheckin.objects.filter(submission_level_category=context['staff'].staff_category,
                                                       submission_pms=context['pms'])

        if submissions:
            submission_checkin = submissions.first()
            context['April'] = submission_checkin.submission_april_checkin_override
            context['May'] = submission_checkin.submission_may_checkin_override
            context['June'] = submission_checkin.submission_june_checkin_override
            context['July'] = submission_checkin.submission_july_checkin_override
            context['August'] = submission_checkin.submission_august_checkin_override
            context['September'] = submission_checkin.submission_september_checkin_override
            context['October'] = submission_checkin.submission_october_checkin_override
            context['November'] =submission_checkin.submission_november_checkin_override
            context['December'] = submission_checkin.submission_december_checkin_override
            context['January'] = submission_checkin.submission_january_checkin_override
            context['February'] = submission_checkin.submission_february_checkin_override
            context['March'] = submission_checkin.submission_march_checkin_override

        context['april_check_in'] = False
        context['may_check_in'] = False
        context['june_check_in'] = False
        context['july_check_in'] = False
        context['august_check_in'] = False
        context['september_check_in'] = False
        context['october_check_in'] = False
        context['november_check_in'] = False
        context['december_check_in'] = False
        context['january_check_in'] = False
        context['february_check_in'] = False
        context['march_check_in'] = False

        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month="April"):
            context['april_check_in'] = True
        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month="May"):
            context['may_check_in'] = True
        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month="June"):
            context['june_check_in'] = True
        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month="July"):
            context['july_check_in'] = True
        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month="August"):
            context['august_check_in'] = True
        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month="September"):
            context['september_check_in'] = True
        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month="October"):
            context['october_check_in'] = True
        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month="November"):
            context['november_check_in'] = True
        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month="December"):
            context['december_check_in'] = True
        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month="January"):
            context['january_check_in'] = True
        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month="February"):
            context['february_check_in'] = True
        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month="March"):
            context['march_check_in'] = True

        return context


@method_decorator(login_required, name='dispatch')
class MyCheckInCreatePrevious(CreateView):
    form_class = CheckInForm

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(MyCheckInCreatePrevious, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)

        if CheckIn.objects.filter(check_in_Staff=context['staff'], check_in_month=self.kwargs['month']):
            context['done'] = True
        else:
            context['done'] = False

        submissions = SubmissionCheckin.objects.filter(submission_pms=context['pms'],
                                                       submission_level_category=context['staff'].staff_category)

        context['allowed'] = False
        if submissions:
            submission_checkin = submissions.first()

            if self.kwargs['month'] == 'April' and submission_checkin.submission_april_checkin_override is True:
                context['allowed'] = True
            if self.kwargs['month'] == 'May' and submission_checkin.submission_may_checkin_override is True:
                context['allowed'] = True
            if self.kwargs['month'] == 'June' and submission_checkin.submission_june_checkin_override is True:
                context['allowed'] = True
            if self.kwargs['month'] == 'July' and submission_checkin.submission_july_checkin_override is True:
                context['allowed'] = True
            if self.kwargs['month'] == 'August' and submission_checkin.submission_august_checkin_override is True:
                context['allowed'] = True
            if self.kwargs['month'] == 'September' and submission_checkin.submission_september_checkin_override is True:
                context['allowed'] = True
            if self.kwargs['month'] == 'October' and submission_checkin.submission_october_checkin_override is True:
                context['allowed'] = True
            if self.kwargs['month'] == 'November' and submission_checkin.submission_november_checkin_override is True:
                context['allowed'] = True
            if self.kwargs['month'] == 'December' and submission_checkin.submission_december_checkin_override is True:
                context['allowed'] = True
            if self.kwargs['month'] == 'January' and submission_checkin.submission_january_checkin_override is True:
                context['allowed'] = True
            if self.kwargs['month'] == 'February' and submission_checkin.submission_february_checkin_override is True:
                context['allowed'] = True
            if self.kwargs['month'] == 'March' and submission_checkin.submission_march_checkin_override is True:
                context['allowed'] = True

        context['month'] = self.kwargs['month']
        context['CheckIn'] = CheckIn.objects.filter(check_in_Staff=context['staff'])
        return context

    def get_success_url(self):
        return '{}'.format(reverse('Site:My_CheckIn_Create_Previous', kwargs={'company_id': self.kwargs['company_id']}))

    def get_initial(self):
        initial = super(MyCheckInCreatePrevious, self).get_initial()
        context = {}
        global_context(self.kwargs['company_id'], self.request.user, context)

        initial['check_in_pms'] = context['pms']
        initial['check_in_Staff'] = context['staff']
        initial['check_in_submit_date'] = datetime.datetime.now()
        initial['check_in_month'] = self.kwargs['month']
        return initial

    def form_valid(self, form):
        super(MyCheckInCreatePrevious, self).form_valid(form)
        staff = get_staff_account(get_company(self.kwargs['company_id']), self.request.user)

        notification_log("CheckIn", "None", staff.staff_person.get_full_name(), staff.staff_person.email,
                         "My CheckIn Submit", "It appears you have submitted a Check-In")
        if get_staff_level(staff) is not None:
            level = get_staff_level(staff)
            notification_log("CheckIn", "None", level.level_head.staff_person.get_full_name(),
                             level.level_head.staff_person.email, "Staff CheckIn Submitted",
                             "It appears " + staff.staff_person.get_full_name() +
                             " has submitted a checkin for your review")

        return HttpResponseRedirect(reverse('Site:My_CheckIn_Create_Previous',
                                            kwargs={'company_id': self.kwargs['company_id']}))


@method_decorator(login_required, name='dispatch')
class MyCheckInView(DetailView):
    model = CheckIn
    context_object_name = 'CheckIn'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(MyCheckInView, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        return context


@method_decorator(login_required, name='dispatch')
class MyCheckInEdit(UpdateView):
    model = CheckIn
    form_class = CheckInForm
    context_object_name = 'CheckIn'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(MyCheckInEdit, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        return context

    def get_success_url(self):
        return '{}'.format(reverse('Site:My_CheckIn_View', kwargs={'company_id': self.kwargs['company_id'],
                                                                   'pk': self.kwargs['pk']}))

    def form_valid(self, form):
        super(MyCheckInEdit, self).form_valid(form)
        staff = get_staff_account(get_company(self.kwargs['company_id']), self.request.user)
        checkin = get_object_or_404(CheckIn, check_in_id=self.kwargs['pk'])
        notification_log("CheckIn", self.kwargs['pk'], staff.staff_person.get_full_name(), staff.staff_person.email,
                         "My Check-In: " + checkin.check_in_month,
                         "It appears you have modified check-in for month " + checkin.check_in_month)
        if get_staff_level(staff) is not None:
            level = get_staff_level(staff)
            notification_log("CheckIn", self.kwargs['pk'], level.level_head.staff_person.get_full_name(),
                             level.level_head.staff_person.email,
                             "Staff Check-In: " + checkin.check_in_month,
                             "It appears " + staff.staff_person.get_full_name() +
                             " has modified or check-in for month  " + checkin.check_in_month +
                             " for your review")

        return HttpResponseRedirect(reverse('Site:My_CheckIn_View', kwargs={'company_id': self.kwargs['company_id'],
                                                                            'pk': self.kwargs['pk']}))


@method_decorator(login_required, name='dispatch')
class MyCheckInDelete(DeleteView):
    model = CheckIn
    context_object_name = 'CheckIn'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(MyCheckInDelete, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)

        return context

    def get_success_url(self):
        return '{}'.format(reverse('Site:My_CheckIn', kwargs={'company_id': self.kwargs['company_id']}))


@method_decorator(login_required, name='dispatch')
class CheckInLevelDown(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(CheckInLevelDown, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        context['level_head'] = Level.objects.filter(level_head=context['staff'],
                                                     level_category__category_company__company_id=context['company_id'])
        levels_down = []
        for level in context['level_head']:
            all_levels_down(level, levels_down)
        context['levels_down'] = levels_down

        return context


@method_decorator(login_required, name='dispatch')
class CheckInLevelDownDetail(DetailView):
    model = Level

    def get_context_data(self, **kwargs):
        context = super(CheckInLevelDownDetail, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)

        level = get_object_or_404(Level, level_id=self.kwargs['pk'])
        context['staff_members'] = LevelMembership.objects.filter(membership_level=level)
        context['staff_members_active'] = LevelMembership.objects.filter(membership_is_active=True,
                                                                         membership_level=level)
        context['staff_members_inactive'] = LevelMembership.objects.filter(membership_is_active=False,
                                                                           membership_level=level)

        members = []
        for member in LevelMembership.objects.filter(membership_level=level):
            kpi = CheckIn.objects.filter(check_in_Staff=member.membership_staff, check_in_pms=context['pms'])

            pending_ci = kpi.filter(check_in_status="Pending").count()
            approved_ci = kpi.filter(check_in_status="Approved").count()

            members.append([member.membership_staff, pending_ci, approved_ci,
                            calculate_overall_check_in_score(member.membership_staff, context['pms'])])
        context['members'] = members
        context['type_id'] = self.kwargs['type_id']

        return context


@method_decorator(login_required, name='dispatch')
class CheckInLevelDownDetailStaff(DetailView):
    model = Staff

    def get_context_data(self, **kwargs):
        context = super(CheckInLevelDownDetailStaff, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)

        context['level'] = get_object_or_404(Level, level_id=self.kwargs['lev_id'])
        context['select_staff'] = get_object_or_404(Staff, staff_id=self.kwargs['pk'])

        context['cis'] = CheckIn.objects.filter(check_in_Staff=context['select_staff'], check_in_pms=context['pms'])
        context['type_id'] = self.kwargs['type_id']

        return context


def staff_check_in_approve(request, company_id, lev_id, type_id, staff_id, pk):
    if request.method == "POST":

        def get_value(x):
            return str(x).strip()

        check_in_id = get_value(request.POST.get("check_in_id"))
        check_in_comments = get_value(request.POST.get("check_in_comments"))

        if int(pk) == int(check_in_id):
            checkin = get_object_or_404(CheckIn, check_in_id=check_in_id)
            checkin.check_in_approver = get_staff_account(get_company(company_id), request.user)
            checkin.check_in_approval_date = datetime.datetime.now()
            checkin.check_in_status = "Approved"
            checkin.check_in_team_leader_comment = check_in_comments
            checkin.save()

            staff_member = get_object_or_404(Staff, staff_id=staff_id)
            team_leader = get_staff_account(get_company(company_id), request.user)

            checkin = get_object_or_404(CheckIn, check_in_id=pk)
            notification_log("CheckIn", pk, team_leader.staff_person.get_full_name(), team_leader.staff_person.email,
                             "Staff Check-In: " + checkin.check_in_month,
                             "It appears you have approved check-in for staff " +
                             staff_member.staff_person.get_full_name() + " for the month " + checkin.check_in_month)

            notification_log("CheckIn", pk, staff_member.staff_person.get_full_name(),
                             staff_member.staff_person.email,
                             "My Check-In: " + checkin.check_in_month,
                             "It appears checkin for the month " + checkin.check_in_month +
                             " has been approved by " + team_leader.staff_person.get_full_name())

    return HttpResponseRedirect(reverse('Site:Check_in_LevelDownDetailStaff',
                                        kwargs={'company_id': company_id, 'lev_id': lev_id, 'type_id': type_id,
                                                'pk': staff_id}))


@method_decorator(login_required, name='dispatch')
class AssessmentList(TemplateView):
    model = Assessment
    context_object_name = 'Assessment'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(AssessmentList, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)

        assessment = Assessment.objects.filter(assessment_pms=context['pms'])
        context['assessment'] = assessment
        context['assessment_current'] = assessment.filter(assessment_start_date__lte=datetime.datetime.now(),
                                                          assessment_end_date__gte=datetime.datetime.now())
        context['assessment_future'] = assessment.filter(assessment_start_date__gt=datetime.datetime.now())
        context['assessment_past'] = assessment.filter(assessment_end_date__lt=datetime.datetime.now())

        return context


@method_decorator(login_required, name='dispatch')
class AssessmentView(DetailView):
    model = Assessment
    context_object_name = 'Assessment'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(AssessmentView, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        assessment = get_object_or_404(Assessment, assessment_id=self.kwargs['pk'])

        if assessment.assessment_start_date <= datetime.datetime.now(
                assessment.assessment_start_date.tzinfo) <= assessment.assessment_end_date:
            context['assessment_status'] = True
        else:
            context['assessment_status'] = False

        all_questions = Questions.objects.filter(question_assessment=assessment)
        context['questions'] = all_questions
        context['top_questions'] = all_questions.filter(question_direction='Top')
        context['bottom_questions'] = all_questions.filter(question_direction='Bottom')

        level_members = []
        level_heads = []

        # Get subordinate staff
        if Level.objects.filter(level_head=context['staff']):
            for level in Level.objects.filter(level_head=context['staff']):
                if LevelMembership.objects.filter(membership_level=level, membership_is_active=True):
                    for records in LevelMembership.objects.filter(membership_level=level, membership_is_active=True):
                        count = 0
                        for question in context['bottom_questions']:
                            if QuestionResponses.objects.filter(response_question=question,
                                                                response_evaluated=records.membership_staff,
                                                                response_staff=context['staff']):
                                count += 1
                        status = str(count) + "/" + str(context['bottom_questions'].count())
                        level_members.append([records, status])

        # Get Team Leaders
        if LevelMembership.objects.filter(membership_staff=context['staff'], membership_is_active=True):
            for level in LevelMembership.objects.filter(membership_staff=context['staff'], membership_is_active=True):
                count = 0
                for question in context['top_questions']:
                    if QuestionResponses.objects.filter(response_question=question,
                                                        response_evaluated=level.membership_level.level_head,
                                                        response_staff=context['staff']):
                        count += 1
                status = str(count) + "/" + str(context['top_questions'].count())
                level_heads.append([level.membership_level, status])

        context['level_heads'] = level_heads
        context['level_members'] = level_members

        return context


@method_decorator(login_required, name='dispatch')
class AssessmentViewMember(DetailView):
    model = Assessment
    context_object_name = 'Assessment'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(AssessmentViewMember, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        assessment = get_object_or_404(Assessment, assessment_id=self.kwargs['pk'])
        member = get_object_or_404(Staff, staff_id=self.kwargs['uid'])
        if assessment.assessment_start_date <= datetime.datetime.now(
                assessment.assessment_start_date.tzinfo) <= assessment.assessment_end_date:
            context['assessment_status'] = True
        else:
            context['assessment_status'] = False

        question_responses = []
        questions = Questions.objects.filter(question_assessment=assessment, question_direction=self.kwargs['a_dir'])
        completed = 0

        if questions.count():
            for question in questions:
                response = QuestionResponses.objects.filter(response_question=question, response_staff=context['staff'],
                                                            response_evaluated=member)
                if response:
                    question_responses.append([question, response, 'Done'])
                    completed = completed + 1
                else:
                    question_responses.append([question, None, 'Not Done'])
            completed = completed / questions.count()
        else:
            completed = 0

        context['question_responses'] = question_responses
        context['completed'] = completed
        context['member'] = member
        context['direction'] = self.kwargs['a_dir']

        return context


@method_decorator(login_required, name='dispatch')
class AssessmentViewMemberResponseCreate(CreateView):
    form_class = QuestionResponseForm

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(AssessmentViewMemberResponseCreate, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)

        assessment = get_object_or_404(Assessment, assessment_id=self.kwargs['pk'])
        member = get_object_or_404(Staff, staff_id=self.kwargs['uid'])
        # direction = self.kwargs['a_dir']
        question = get_object_or_404(Questions, question_id=self.kwargs['qid'])

        if assessment.assessment_start_date <= datetime.datetime.now(
                assessment.assessment_start_date.tzinfo) <= assessment.assessment_end_date:
            context['assessment_status'] = True
        else:
            context['assessment_status'] = False

        response = QuestionResponses.objects.filter(response_question_id=question.question_id,
                                                    response_staff=context['staff'], response_evaluated=member)
        if response:
            completed = True
        else:
            completed = False

        context['Assessment'] = assessment
        context['completed'] = completed
        context['question'] = question
        context['member'] = member
        context['direction'] = self.kwargs['a_dir']

        return context

    def get_initial(self):
        initial = super(AssessmentViewMemberResponseCreate, self).get_initial()
        initial['response_staff'] = get_staff_account(get_company(self.kwargs['company_id']), self.request.user)
        initial['response_evaluated'] = get_object_or_404(Staff, staff_id=self.kwargs['uid'])
        initial['response_question'] = get_object_or_404(Questions, question_id=self.kwargs['qid'])
        initial['response_submitted_date'] = datetime.datetime.now()

        return initial

    def get_success_url(self):
        return '{}'.format(reverse('Site:Assessment_View_Member',
                                   kwargs={'company_id': self.kwargs['company_id'], 'pk': self.kwargs['pk'],
                                           'uid': self.kwargs['uid'], 'a_dir': self.kwargs['a_dir']}))


@method_decorator(login_required, name='dispatch')
class AssessmentViewMemberResponseEdit(UpdateView):
    pk_url_kwarg = 'qrid'
    model = QuestionResponses
    form_class = QuestionResponseForm

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(AssessmentViewMemberResponseEdit, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)

        assessment = get_object_or_404(Assessment, assessment_id=self.kwargs['pk'])
        member = get_object_or_404(Staff, staff_id=self.kwargs['uid'])
        # response = get_object_or_404(QuestionResponses, response_id=self.kwargs['qrid'])

        if assessment.assessment_start_date <= datetime.datetime.now(assessment.assessment_start_date.tzinfo) <= \
                assessment.assessment_end_date:
            context['assessment_status'] = True
        else:
            context['assessment_status'] = False

        context['Assessment'] = assessment
        context['member'] = member
        context['direction'] = self.kwargs['a_dir']

        return context

    def get_success_url(self):
        return '{}'.format(reverse('Site:Assessment_View_Member', kwargs={'company_id': self.kwargs[''],
                                                                          'pk': self.kwargs['pk'],
                                                                          'uid': self.kwargs['uid'],
                                                                          'a_dir': self.kwargs['a_dir']}))


@method_decorator(login_required, name='dispatch')
class AssessmentViewMy(DetailView):
    model = Assessment
    context_object_name = 'Assessment'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(AssessmentViewMy, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        assessment = get_object_or_404(Assessment, assessment_id=self.kwargs['pk'])
        if datetime.datetime.now(assessment.assessment_start_date.tzinfo) > assessment.assessment_end_date:
            context['assessment_status'] = True
        else:
            context['assessment_status'] = False

        question_responses = []
        questions = Questions.objects.filter(question_assessment=assessment, question_direction=self.kwargs['a_dir'])

        if questions.count():
            for question in questions:
                score = 0
                response = QuestionResponses.objects.filter(response_question=question,
                                                            response_evaluated=context['staff'])
                comments = []
                if response:
                    for res in response:
                        score = score + res.response_submitted
                        comments.append(res.response_comment)
                    score = round(score / response.count(), 2)
                    question_responses.append([question, response, 'Done', score, comments])
                else:
                    question_responses.append([question, None, 'Not Done', score, comments])

        context['question_responses'] = question_responses
        context['direction'] = self.kwargs['a_dir']
        return context


def staff_create_question_response(request, company_id, aid, staff_id, a_dir, qid):
    if request.method == "POST":

        def get_value(x):
            if str(x).strip():
                return str(x).strip()
            else:
                return None

        def get_value_submitted(x):
            if float(x):
                return float(x)
            else:
                return None

        response = QuestionResponses()
        response.response_submitted_date = datetime.datetime.now()
        response.response_staff = get_staff_account(get_company(company_id), request.user)
        response.response_evaluated = get_object_or_404(Staff, staff_id=staff_id)
        response.response_question = get_object_or_404(Questions, question_id=qid)
        response.response_comment = get_value(request.POST.get("question_response_comment"))
        response.response_submitted = get_value_submitted(request.POST.get("question_response_value"))
        response.save()

    return HttpResponseRedirect(reverse('Site:Assessment_View_Member',
                                        kwargs={'company_id': company_id, 'pk': aid, 'uid': staff_id, 'a_dir': a_dir}))


def staff_edit_question_response(request, company_id, aid, staff_id, a_dir, rid):
    if request.method == "POST":

        def get_value(x):
            if str(x).strip():
                return str(x).strip()
            else:
                return None

        def get_value_submitted(x):
            if float(x):
                return float(x)
            else:
                return None

        res_id = int(request.POST.get("response_id"))

        if res_id == rid:
            response = get_object_or_404(QuestionResponses, response_id=rid)
            response.response_comment = get_value(request.POST.get("question_response_comment"))
            response.response_submitted = get_value_submitted(request.POST.get("question_response_value"))
            response.save()

    return HttpResponseRedirect(reverse('Site:Assessment_View_Member',
                                        kwargs={'company_id': company_id, 'pk': aid, 'uid': staff_id, 'a_dir': a_dir}))


@method_decorator(login_required, name='dispatch')
class Report(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(Report, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        staff = context['staff']

        my_members_score = []
        member_error_list = []
        levels_down_score = []

        average_score = 0
        cat_75 = 0
        cat_50 = 0
        cat_25 = 0
        cat_0 = 0

        if staff.staff_superuser:
            for member_staff in Staff.objects.filter(staff_company=context['company']):
                level = get_staff_level(member_staff)
                assessment_score = calculate_overall_assessment_score(member_staff, context['pms'])
                check_in_score = calculate_overall_check_in_score(member_staff, context['pms'])
                bu_score = get_bu_score(member_staff, context['pms'])
                company_score = get_company_score(member_staff, context['pms'])
                the_bu = get_bu(member_staff)
                the_company = get_company_used(member_staff)

                try:
                    kpi_score = calculate_overall_kpi_score(member_staff, context['pms'])
                    overall_score = calculate_overall_score(member_staff, context['pms'])
                except Exception as e:
                    kpi_score = 0
                    overall_score = 0
                    member_error_list.append([member_staff, e])

                average_score += overall_score
                if overall_score >= 75:
                    cat_75 += 1
                elif 50 <= overall_score < 75:
                    cat_50 += 1
                elif 25 <= overall_score < 50:
                    cat_25 += 1
                else:
                    cat_0 += 1

                my_members_score.append([member_staff, level, assessment_score, check_in_score, kpi_score, bu_score,
                                         company_score, overall_score, the_bu, the_company])
            if Staff.objects.filter(staff_company=context['company']).count() > 0:
                average_score = round(average_score/Staff.objects.filter(staff_company=context['company']).count(), 2)

        else:
            for level in Level.objects.filter(level_head=staff):
                level_score = 0
                for members in LevelMembership.objects.filter(membership_level=level):
                    level = members.membership_level
                    assessment_score = calculate_overall_assessment_score(members.membership_staff, context['pms'])
                    check_in_score = calculate_overall_check_in_score(members.membership_staff, context['pms'])
                    kpi_score = calculate_overall_kpi_score(members.membership_staff, context['pms'])
                    bu_score = get_bu_score(members.membership_staff, context['pms'])
                    company_score = get_company_score(members.membership_staff, context['pms'])
                    overall_score = calculate_overall_score(members.membership_staff, context['pms'])
                    the_bu = get_bu(members)
                    the_company = get_company_used(members)

                    level_score += overall_score
                    if overall_score >= 75:
                        cat_75 += 1
                    elif 50 <= overall_score < 75:
                        cat_50 += 1
                    elif 25 <= overall_score < 50:
                        cat_25 += 1
                    else:
                        cat_0 += 1

                    my_members_score.append([members.membership_staff, level, assessment_score, check_in_score,
                                             kpi_score, bu_score, company_score, overall_score, the_bu, the_company])
                if LevelMembership.objects.filter(membership_level=level).count() > 0:
                    level_score = round(level_score/LevelMembership.objects.filter(membership_level=level).count(), 2)

                average_score += level_score

            if Level.objects.filter(level_head=staff).count() > 0:
                average_score = round(average_score/Level.objects.filter(level_head=staff).count(), 2)

            if check_staff_is_level_head(self.kwargs['company_id'], staff):
                levels_down = []
                for levels in Level.objects.filter(level_head=staff):
                    all_levels_down(levels, levels_down)
                if len(levels_down) > 0:
                    average_score = 0
                for level in levels_down:
                    level_score = 0
                    for members in LevelMembership.objects.filter(membership_level=level):
                        level = members.membership_level
                        assessment_score = calculate_overall_assessment_score(members.membership_staff, context['pms'])
                        check_in_score = calculate_overall_check_in_score(members.membership_staff, context['pms'])
                        kpi_score = calculate_overall_kpi_score(members.membership_staff, context['pms'])
                        bu_score = get_bu_score(members.membership_staff, context['pms'])
                        company_score = get_company_score(members.membership_staff, context['pms'])
                        overall_score = calculate_overall_score(members.membership_staff, context['pms'])
                        the_bu = get_bu(members)
                        the_company = get_company_used(members)

                        level_score += overall_score
                        if overall_score >= 75:
                            cat_75 += 1
                        elif 50 <= overall_score < 75:
                            cat_50 += 1
                        elif 25 <= overall_score < 50:
                            cat_25 += 1
                        else:
                            cat_0 += 1

                        levels_down_score.append([members.membership_staff, level, assessment_score, check_in_score,
                                                  kpi_score, bu_score, company_score, overall_score, the_bu, the_company])
                    if LevelMembership.objects.filter(membership_level=level).count() > 0:
                        level_score = round(level_score/LevelMembership.objects.filter(membership_level=level).count(),
                                            2)
                    average_score += level_score
                if len(levels_down) > 0:
                    average_score = round(average_score/len(levels_down), 2)

        context['my_members_score'] = my_members_score
        context['levels_down_score'] = levels_down_score
        context['cat_75'] = cat_75
        context['cat_50'] = cat_50
        context['cat_25'] = cat_25
        context['cat_0'] = cat_0
        context['average_score'] = average_score
        return context


@method_decorator(login_required, name='dispatch')
class ReportKPIResults(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(ReportKPIResults, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        staff = context['staff']

        my_members_kpi_score = []
        member_error_list = []
        levels_down_kpi_score = []
        average_score = 0
        cat_75 = 0
        cat_50 = 0
        cat_25 = 0
        cat_0 = 0

        if staff.staff_superuser:
            for member_staff in Staff.objects.filter(staff_company=context['company']):
                kpis = []
                kpi_type = KPIType.objects.filter(type_pms=context['pms'],
                                                  type_category=member_staff.staff_category)
                if kpi_type:
                    k_type = kpi_type.first().type_kpi
                else:
                    k_type = "Annual Target"
                for kpi in KPI.objects.filter(kpi_staff=member_staff, kpi_pms=context['pms']):
                    kpis.append([kpi, calculate_kpi_score(kpi, k_type)])

                try:
                    overall_kpi_score = calculate_overall_kpi_score(member_staff, context['pms'])
                except Exception as e:
                    overall_kpi_score = 0
                    member_error_list.append([member_staff, e])
                average_score += overall_kpi_score
                if overall_kpi_score >= 75:
                    cat_75 += 1
                elif 50 <= overall_kpi_score < 75:
                    cat_50 += 1
                elif 25 <= overall_kpi_score < 50:
                    cat_25 += 1
                else:
                    cat_0 += 1
                bu = get_bu(member_staff)
                my_members_kpi_score.append([member_staff, kpis,
                                             calculate_overall_kpi_score(member_staff, context['pms']), kpis,
                                             bu, get_staff_level(member_staff)])
            if Staff.objects.filter(staff_company=context['company']).count() > 0:
                average_score = round(average_score/Staff.objects.filter(staff_company=context['company']).count(), 2)

        else:
            for level in Level.objects.filter(level_head=staff):
                kpi_score = 0
                for members in LevelMembership.objects.filter(membership_level=level):
                    kpis = []
                    kpi_type = KPIType.objects.filter(type_pms=context['pms'],
                                                      type_category=members.membership_staff.staff_category)
                    if kpi_type:
                        k_type = kpi_type.first().type_kpi
                    else:
                        k_type = "Annual Target"
                    for kpi in KPI.objects.filter(kpi_staff=members.membership_staff, kpi_pms=context['pms']):
                        kpis.append([kpi, calculate_kpi_score(kpi, k_type)])

                    overall_kpi_score = calculate_overall_kpi_score(members.membership_staff, context['pms'])
                    kpi_score += overall_kpi_score
                    if overall_kpi_score >= 75:
                        cat_75 += 1
                    elif 50 <= overall_kpi_score < 75:
                        cat_50 += 1
                    elif 25 <= overall_kpi_score < 50:
                        cat_25 += 1
                    else:
                        cat_0 += 1
                    bu = get_bu(members)

                    my_members_kpi_score.append([members.membership_staff, kpis,
                                                 calculate_overall_kpi_score(members.membership_staff, context['pms']),
                                                 bu, get_staff_level(members)],)
                if LevelMembership.objects.filter(membership_level=level).count() > 0:
                    kpi_score = round(kpi_score/LevelMembership.objects.filter(membership_level=level).count(), 2)
                average_score += kpi_score

            if Level.objects.filter(level_head=staff).count() > 0:
                average_score += round(average_score/Level.objects.filter(level_head=staff).count(), 2)

            if check_staff_is_level_head(self.kwargs['company_id'], staff):
                levels_down = []
                for levels in Level.objects.filter(level_head=staff):
                    all_levels_down(levels, levels_down)

                if len(levels_down) > 0:
                    average_score = 0

                for level in levels_down:
                    kpi_score = 0
                    for members in LevelMembership.objects.filter(membership_level=level):
                        kpis = []
                        kpi_type = KPIType.objects.filter(type_pms=context['pms'],
                                                          type_category=members.membership_staff.staff_category)
                        if kpi_type:
                            k_type = kpi_type.first().type_kpi
                        else:
                            k_type = "Annual Target"
                        for kpi in KPI.objects.filter(kpi_staff=members.membership_staff, kpi_pms=context['pms']):
                            kpis.append([kpi, calculate_kpi_score(kpi, k_type)])

                        overall_kpi_score = calculate_overall_kpi_score(members.membership_staff, context['pms'])
                        kpi_score += overall_kpi_score
                        if overall_kpi_score >= 75:
                            cat_75 += 1
                        elif 50 <= overall_kpi_score < 75:
                            cat_50 += 1
                        elif 25 <= overall_kpi_score < 50:
                            cat_25 += 1
                        else:
                            cat_0 += 1
                        bu = get_bu(members)
                        levels_down_kpi_score.append([members.membership_staff, kpis,
                                                      calculate_overall_kpi_score(members.membership_staff, context['pms']),
                                                      bu, get_staff_level(members)])
                    if LevelMembership.objects.filter(membership_level=level).count() > 0:
                        kpi_score = round(kpi_score / LevelMembership.objects.filter(membership_level=level).count(), 2)
                    average_score += kpi_score
                if len(levels_down) > 0:
                    average_score = round(average_score/len(levels_down), 2)

        context['cat_75'] = cat_75
        context['cat_50'] = cat_50
        context['cat_25'] = cat_25
        context['cat_0'] = cat_0
        context['average_score'] = average_score
        context['my_members_kpi_score'] = my_members_kpi_score
        context['levels_down_kpi_score'] = levels_down_kpi_score
        return context


@method_decorator(login_required, name='dispatch')
class ReportKPISubmissions(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(ReportKPISubmissions, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        staff = context['staff']

        my_members_submissions = []
        levels_down_submissions = []

        staff_with_pending = 0
        staff_with_required = 0
        staff_with_less = 0
        minimum_kpi = 5

        if staff.staff_superuser:
            for member_staff in Staff.objects.filter(staff_company=context['company']):
                kpis = KPI.objects.filter(kpi_staff=member_staff, kpi_pms=context['pms'])
                submitted = kpis.filter(kpi_status="Submitted")
                approved = kpis.filter(kpi_status="Approved")
                pending = kpis.filter(kpi_status="Pending")
                considered = submitted.count() + approved.count() + pending.count()

                if pending.count() > 0:
                    staff_with_pending += 1

                submission = SubmissionKPI.objects.filter(submission_level_category=member_staff.staff_category)
                if submission:
                    minimum_kpi = submission.first().submission_minimum_number

                if considered >= minimum_kpi:
                    staff_with_required += 1
                else:
                    staff_with_less += 1

                my_members_submissions.append([member_staff, submitted.count(), pending.count(), approved.count(),
                                               considered, minimum_kpi])

        else:
            for level in Level.objects.filter(level_head=staff):
                for members in LevelMembership.objects.filter(membership_level=level):
                    kpis = KPI.objects.filter(kpi_staff=members.membership_staff, kpi_pms=context['pms'])
                    submitted = kpis.filter(kpi_status="Submitted")
                    approved = kpis.filter(kpi_status="Approved")
                    pending = kpis.filter(kpi_status="Pending")
                    considered = submitted.count() + approved.count() + pending.count()

                    if pending.count() > 0:
                        staff_with_pending += 1

                    submission = SubmissionKPI.objects.filter(
                        submission_level_category=members.membership_staff.staff_category)
                    if submission:
                        minimum_kpi = submission.first().submission_minimum_number

                    if considered >= minimum_kpi:
                        staff_with_required += 1
                    else:
                        staff_with_less += 1

                    my_members_submissions.append([members.membership_staff, submitted.count(), pending.count(),
                                                   approved.count(), considered, minimum_kpi])

        if check_staff_is_level_head(self.kwargs['company_id'], staff):
            levels_down = []
            for levels in Level.objects.filter(level_head=staff):
                all_levels_down(levels, levels_down)

            for level in levels_down:
                for members in LevelMembership.objects.filter(membership_level=level):
                    kpis = KPI.objects.filter(kpi_staff=members.membership_staff, kpi_pms=context['pms'])
                    submitted = kpis.filter(kpi_status="Submitted")
                    approved = kpis.filter(kpi_status="Approved")
                    pending = kpis.filter(kpi_status="Pending")
                    considered = submitted.count() + approved.count() + pending.count()

                    if pending.count() > 0:
                        staff_with_pending += 1

                    submission = SubmissionKPI.objects.filter(
                        submission_level_category=members.membership_staff.staff_category)
                    if submission:
                        minimum_kpi = submission.first().submission_minimum_number

                    if considered >= minimum_kpi:
                        staff_with_required += 1
                    else:
                        staff_with_less += 1

                    my_members_submissions.append([members.membership_staff, submitted.count(), pending.count(),
                                                   approved.count(), considered, minimum_kpi])

        context['staff_with_pending'] = staff_with_pending
        context['staff_with_required'] = staff_with_required
        context['staff_with_less'] = staff_with_less

        context['my_members_submission'] = my_members_submissions
        context['levels_down_submission'] = levels_down_submissions
        return context


@method_decorator(login_required, name='dispatch')
class ReportKPISubmissionsResults(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(ReportKPISubmissionsResults, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        staff = context['staff']

        my_members_kpi_score = []

        if staff.staff_superuser:
            for member_staff in Staff.objects.filter(staff_company=context['company']):
                kpis = KPI.objects.filter(kpi_staff=member_staff, kpi_pms=context['pms'])
                approved_kpis = kpis.filter(kpi_status="Approved")
                pending_kpis = kpis.filter(kpi_status="Pending")
                rejected_kpis = kpis.filter(kpi_status="Rejected")
                the_bu = get_bu(member_staff)
                the_company = get_company_used(member_staff)

                apr_r = may_r = jun_r = jul_r = aug_r = sep_r = oct_r = nov_r = dec_r = jan_r = feb_r = mar_r = 0
                apr_a = may_a = jun_a = jul_a = aug_a = sep_a = oct_a = nov_a = dec_a = jan_a = feb_a = mar_a = 0
                
                for kpi in approved_kpis:
                    if kpi.kpi_april_score is not None:
                        apr_r += 1
                    if kpi.kpi_may_score is not None:
                        may_r += 1
                    if kpi.kpi_june_score is not None:
                        jun_r += 1
                    if kpi.kpi_july_score is not None:
                        jul_r += 1
                    if kpi.kpi_august_score is not None:
                        aug_r += 1
                    if kpi.kpi_september_score is not None:
                        sep_r += 1
                    if kpi.kpi_october_score is not None:
                        oct_r += 1
                    if kpi.kpi_november_score is not None:
                        nov_r += 1
                    if kpi.kpi_december_score is not None:
                        dec_r += 1
                    if kpi.kpi_january_score is not None:
                        jan_r += 1
                    if kpi.kpi_february_score is not None:
                        feb_r += 1
                    if kpi.kpi_march_score is not None:
                        mar_r += 1

                    if kpi.kpi_april_score_approve is True:
                        apr_a += 1
                    if kpi.kpi_may_score_approve is True:
                        may_a += 1
                    if kpi.kpi_june_score_approve is True:
                        jun_a += 1
                    if kpi.kpi_july_score_approve is True:
                        jul_a += 1
                    if kpi.kpi_august_score_approve is True:
                        aug_a += 1
                    if kpi.kpi_september_score_approve is True:
                        sep_a += 1
                    if kpi.kpi_october_score_approve is True:
                        oct_a += 1
                    if kpi.kpi_november_score_approve is True:
                        nov_a += 1
                    if kpi.kpi_december_score_approve is True:
                        dec_a += 1
                    if kpi.kpi_january_score_approve is True:
                        jan_a += 1
                    if kpi.kpi_february_score_approve is True:
                        feb_a += 1
                    if kpi.kpi_march_score_approve is True:
                        mar_a += 1
                results = {
                    'apr_r': apr_r, 'may_r': may_r, 'jun_r': jun_r, 'jul_r': jul_r, 'aug_r': aug_r, 'sep_r': sep_r,
                    'oct_r': oct_r, 'nov_r': nov_r, 'dec_r': dec_r, 'jan_r': jan_r, 'feb_r': feb_r, 'mar_r': mar_r
                }

                approvals = {
                    'apr_a': apr_a, 'may_a': may_a, 'jun_a': jun_a, 'jul_a': jul_a, 'aug_a': aug_a, 'sep_a': sep_a,
                    'oct_a': oct_a, 'nov_a': nov_a, 'dec_a': dec_a, 'jan_a': jan_a, 'feb_a': feb_a, 'mar_a': mar_a
                }
                
                my_members_kpi_score.append([member_staff, approved_kpis.count(), pending_kpis.count(),
                                             rejected_kpis.count(), results, approvals, the_bu, the_company],)

        else:
            for level in Level.objects.filter(level_head=staff):
                for members in LevelMembership.objects.filter(membership_level=level):
                    kpis = KPI.objects.filter(kpi_staff=members.membership_staff, kpi_pms=context['pms'])
                    approved_kpis = kpis.filter(kpi_status="Approved")
                    pending_kpis = kpis.filter(kpi_status="Pending")
                    rejected_kpis = kpis.filter(kpi_status="Rejected")
                    the_bu = get_bu(members)
                    the_company = get_company_used(members)

                    apr_r = may_r = jun_r = jul_r = aug_r = sep_r = oct_r = nov_r = dec_r = jan_r = feb_r = mar_r = 0
                    apr_a = may_a = jun_a = jul_a = aug_a = sep_a = oct_a = nov_a = dec_a = jan_a = feb_a = mar_a = 0

                    for kpi in approved_kpis:
                        if kpi.kpi_april_score is not None:
                            apr_r += 1
                        if kpi.kpi_may_score is not None:
                            may_r += 1
                        if kpi.kpi_june_score is not None:
                            jun_r += 1
                        if kpi.kpi_july_score is not None:
                            jul_r += 1
                        if kpi.kpi_august_score is not None:
                            aug_r += 1
                        if kpi.kpi_september_score is not None:
                            sep_r += 1
                        if kpi.kpi_october_score is not None:
                            oct_r += 1
                        if kpi.kpi_november_score is not None:
                            nov_r += 1
                        if kpi.kpi_december_score is not None:
                            dec_r += 1
                        if kpi.kpi_january_score is not None:
                            jan_r += 1
                        if kpi.kpi_february_score is not None:
                            feb_r += 1
                        if kpi.kpi_march_score is not None:
                            mar_r += 1

                        if kpi.kpi_april_score_approve is True:
                            apr_a += 1
                        if kpi.kpi_may_score_approve is True:
                            may_a += 1
                        if kpi.kpi_june_score_approve is True:
                            jun_a += 1
                        if kpi.kpi_july_score_approve is True:
                            jul_a += 1
                        if kpi.kpi_august_score_approve is True:
                            aug_a += 1
                        if kpi.kpi_september_score_approve is True:
                            sep_a += 1
                        if kpi.kpi_october_score_approve is True:
                            oct_a += 1
                        if kpi.kpi_november_score_approve is True:
                            nov_a += 1
                        if kpi.kpi_december_score_approve is True:
                            dec_a += 1
                        if kpi.kpi_january_score_approve is True:
                            jan_a += 1
                        if kpi.kpi_february_score_approve is True:
                            feb_a += 1
                        if kpi.kpi_march_score_approve is True:
                            mar_a += 1
                    results = {
                        'apr_r': apr_r, 'may_r': may_r, 'jun_r': jun_r, 'jul_r': jul_r, 'aug_r': aug_r, 'sep_r': sep_r,
                        'oct_r': oct_r, 'nov_r': nov_r, 'dec_r': dec_r, 'jan_r': jan_r, 'feb_r': feb_r, 'mar_r': mar_r
                    }

                    approvals = {
                        'apr_a': apr_a, 'may_a': may_a, 'jun_a': jun_a, 'jul_a': jul_a, 'aug_a': aug_a, 'sep_a': sep_a,
                        'oct_a': oct_a, 'nov_a': nov_a, 'dec_a': dec_a, 'jan_a': jan_a, 'feb_a': feb_a, 'mar_a': mar_a
                    }

                    my_members_kpi_score.append([members.membership_staff, approved_kpis.count(), pending_kpis.count(),
                                                 rejected_kpis.count(), results, approvals, the_bu, the_company],)


            if check_staff_is_level_head(self.kwargs['company_id'], staff):
                levels_down = []
                for levels in Level.objects.filter(level_head=staff):
                    all_levels_down(levels, levels_down)

                if len(levels_down) > 0:
                    average_score = 0

                for level in levels_down:
                    kpi_score = 0
                    for members in LevelMembership.objects.filter(membership_level=level):
                        kpis = KPI.objects.filter(kpi_staff=members.membership_staff, kpi_pms=context['pms'])
                        kpis = KPI.objects.filter(kpi_staff=members.membership_staff, kpi_pms=context['pms'])
                        approved_kpis = kpis.filter(kpi_status="Approved")
                        pending_kpis = kpis.filter(kpi_status="Pending")
                        rejected_kpis = kpis.filter(kpi_status="Rejected")
                        the_bu = get_bu(members)
                        the_company = get_company_used(members)

                        apr_r = may_r = jun_r = jul_r = aug_r = sep_r = oct_r = nov_r = dec_r = jan_r = feb_r = mar_r = 0
                        apr_a = may_a = jun_a = jul_a = aug_a = sep_a = oct_a = nov_a = dec_a = jan_a = feb_a = mar_a = 0

                        for kpi in approved_kpis:
                            if kpi.kpi_april_score is not None:
                                apr_r += 1
                            if kpi.kpi_may_score is not None:
                                may_r += 1
                            if kpi.kpi_june_score is not None:
                                jun_r += 1
                            if kpi.kpi_july_score is not None:
                                jul_r += 1
                            if kpi.kpi_august_score is not None:
                                aug_r += 1
                            if kpi.kpi_september_score is not None:
                                sep_r += 1
                            if kpi.kpi_october_score is not None:
                                oct_r += 1
                            if kpi.kpi_november_score is not None:
                                nov_r += 1
                            if kpi.kpi_december_score is not None:
                                dec_r += 1
                            if kpi.kpi_january_score is not None:
                                jan_r += 1
                            if kpi.kpi_february_score is not None:
                                feb_r += 1
                            if kpi.kpi_march_score is not None:
                                mar_r += 1

                            if kpi.kpi_april_score_approve is True:
                                apr_a += 1
                            if kpi.kpi_may_score_approve is True:
                                may_a += 1
                            if kpi.kpi_june_score_approve is True:
                                jun_a += 1
                            if kpi.kpi_july_score_approve is True:
                                jul_a += 1
                            if kpi.kpi_august_score_approve is True:
                                aug_a += 1
                            if kpi.kpi_september_score_approve is True:
                                sep_a += 1
                            if kpi.kpi_october_score_approve is True:
                                oct_a += 1
                            if kpi.kpi_november_score_approve is True:
                                nov_a += 1
                            if kpi.kpi_december_score_approve is True:
                                dec_a += 1
                            if kpi.kpi_january_score_approve is True:
                                jan_a += 1
                            if kpi.kpi_february_score_approve is True:
                                feb_a += 1
                            if kpi.kpi_march_score_approve is True:
                                mar_a += 1
                        results = {
                            'apr_r': apr_r, 'may_r': may_r, 'jun_r': jun_r, 'jul_r': jul_r, 'aug_r': aug_r, 'sep_r': sep_r,
                            'oct_r': oct_r, 'nov_r': nov_r, 'dec_r': dec_r, 'jan_r': jan_r, 'feb_r': feb_r, 'mar_r': mar_r
                        }

                        approvals = {
                            'apr_a': apr_a, 'may_a': may_a, 'jun_a': jun_a, 'jul_a': jul_a, 'aug_a': aug_a, 'sep_a': sep_a,
                            'oct_a': oct_a, 'nov_a': nov_a, 'dec_a': dec_a, 'jan_a': jan_a, 'feb_a': feb_a, 'mar_a': mar_a
                        }

                        my_members_kpi_score.append([members.membership_staff, approved_kpis.count(), pending_kpis.count(),
                                                     rejected_kpis.count(), results, approvals, the_bu, the_company],)

        context['my_members_submission'] = my_members_kpi_score
        return context


@method_decorator(login_required, name='dispatch')
class ReportAssessments(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(ReportAssessments, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        staff = context['staff']

        my_members_score = []
        levels_down_score = []

        cat_75 = 0
        cat_50 = 0
        cat_25 = 0
        cat_0 = 0

        context['Assessments'] = Assessment.objects.filter(assessment_pms=context['pms'])

        if staff.staff_superuser:
            for member_staff in Staff.objects.filter(staff_company=context['company']):
                assessment_score = calculate_overall_assessment_score(member_staff, context['pms'])
                my_members_score.append([member_staff, assessment_score])

                if assessment_score >= 75:
                    cat_75 += 1
                elif 50 <= assessment_score < 75:
                    cat_50 += 1
                elif 25 <= assessment_score < 50:
                    cat_25 += 1
                else:
                    cat_0 += 1
        else:
            for level in Level.objects.filter(level_head=staff):
                for members in LevelMembership.objects.filter(membership_level=level):
                    assessment_score = calculate_overall_assessment_score(members.membership_staff, context['pms'])

                    my_members_score.append([members.membership_staff, assessment_score])

                    if assessment_score >= 75:
                        cat_75 += 1
                    elif 50 <= assessment_score < 75:
                        cat_50 += 1
                    elif 25 <= assessment_score < 50:
                        cat_25 += 1
                    else:
                        cat_0 += 1

            if check_staff_is_level_head(self.kwargs['company_id'], staff):
                levels_down = []
                for levels in Level.objects.filter(level_head=staff):
                    all_levels_down(levels, levels_down)

                for level in levels_down:
                    for members in LevelMembership.objects.filter(membership_level=level):
                        assessment_score = calculate_overall_assessment_score(members.membership_staff, context['pms'])

                        levels_down_score.append([members.membership_staff, assessment_score])

                        if assessment_score >= 75:
                            cat_75 += 1
                        elif 50 <= assessment_score < 75:
                            cat_50 += 1
                        elif 25 <= assessment_score < 50:
                            cat_25 += 1
                        else:
                            cat_0 += 1

        context['cat_75'] = cat_75
        context['cat_50'] = cat_50
        context['cat_25'] = cat_25
        context['cat_0'] = cat_0
        context['my_members_score'] = my_members_score
        context['levels_down_score'] = levels_down_score
        return context


@method_decorator(login_required, name='dispatch')
class ReportAssessmentsDetails(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(ReportAssessmentsDetails, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        staff = context['staff']

        my_members_score = []
        levels_down_score = []
        cat_75 = 0
        cat_50 = 0
        cat_25 = 0
        cat_0 = 0


        assessment = get_object_or_404(Assessment, assessment_id=self.kwargs['pk'])

        if staff.staff_superuser:
            for member_staff in Staff.objects.filter(staff_company=context['company']):
                s_tl = calculate_assessment_score(assessment, member_staff, "Top")
                tl_s = calculate_assessment_score(assessment, member_staff, "Bottom")
                score = calculate_one_assessment_score(member_staff, context['pms'], assessment)
                
                if score >= 75:
                    cat_75 += 1
                elif 50 <= score < 75:
                    cat_50 += 1
                elif 25 <= score < 50:
                    cat_25 += 1
                else:
                    cat_0 += 1

                my_members_score.append([member_staff, s_tl, tl_s, score])
        else:
            for level in Level.objects.filter(level_head=staff):
                for members in LevelMembership.objects.filter(membership_level=level):
                    s_tl = calculate_assessment_score(assessment, members.membership_staff, "Top")
                    tl_s = calculate_assessment_score(assessment, members.membership_staff, "Bottom")
                    score = calculate_one_assessment_score(members.membership_staff, context['pms'], assessment)

                    if score >= 75:
                        cat_75 += 1
                    elif 50 <= score < 75:
                        cat_50 += 1
                    elif 25 <= score < 50:
                        cat_25 += 1
                    else:
                        cat_0 += 1

                    my_members_score.append([members.membership_staff, s_tl, tl_s, score])

            if check_staff_is_level_head(self.kwargs['company_id'], staff):
                levels_down = []
                for levels in Level.objects.filter(level_head=staff):
                    all_levels_down(levels, levels_down)

                for level in levels_down:
                    for members in LevelMembership.objects.filter(membership_level=level):
                        s_tl = calculate_assessment_score(assessment, members.membership_staff, "Top")
                        tl_s = calculate_assessment_score(assessment, members.membership_staff, "Bottom")
                        score = calculate_one_assessment_score(members.membership_staff, context['pms'], assessment)

                        if score >= 75:
                            cat_75 += 1
                        elif 50 <= score < 75:
                            cat_50 += 1
                        elif 25 <= score < 50:
                            cat_25 += 1
                        else:
                            cat_0 += 1

                        levels_down_score.append([members.membership_staff, s_tl, tl_s, score])

        context['cat_75'] = cat_75
        context['cat_50'] = cat_50
        context['cat_25'] = cat_25
        context['cat_0'] = cat_0
        context['assessment'] = assessment
        context['my_members_score'] = my_members_score
        context['levels_down_score'] = levels_down_score
        return context


@method_decorator(login_required, name='dispatch')
class ReportAssessmentsParticipation(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(ReportAssessmentsParticipation, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        staff = context['staff']

        s_tl = []
        tl_s = []
        s_tl_participated = 0
        s_tl_not_participated = 0
        s_tl_partially_participated = 0
        
        tl_s_not_participated = 0
        tl_s_participated = 0
        tl_s_partially_participated = 0
        
        assessment = get_object_or_404(Assessment, assessment_id=self.kwargs['pk'])

        all_questions = Questions.objects.filter(question_assessment=assessment)
        context['questions'] = all_questions
        top_questions = all_questions.filter(question_direction='Top')
        bottom_questions = all_questions.filter(question_direction='Bottom')

        if staff.staff_superuser:
            for member_staff in Staff.objects.filter(staff_company=context['company']):
                levels_count = Level.objects.filter(level_head=member_staff).count()
                if Level.objects.filter(level_head=member_staff):
                    for level in Level.objects.filter(level_head=member_staff):
                        if LevelMembership.objects.filter(membership_level=level, membership_is_active=True):
                            for records in LevelMembership.objects.filter(membership_level=level,
                                                                          membership_is_active=True):
                                count = 0
                                for question in bottom_questions:
                                    if QuestionResponses.objects.filter(response_question=question,
                                                                        response_evaluated=records.membership_staff):
                                        count += 1
                                if bottom_questions.count() == 0:
                                    status = 0
                                else:
                                    status = round(count/bottom_questions.count(), 2)
                                
                                if status <= 0:
                                    tl_s_not_participated += 1
                                elif 0 < status < 1:
                                    tl_s_partially_participated += 1
                                else:
                                    tl_s_participated += 1
                
                                tl_s.append([member_staff, records.membership_staff, status])
                                
                if LevelMembership.objects.filter(membership_staff=member_staff, membership_is_active=True):
                    for level in LevelMembership.objects.filter(membership_staff=member_staff,
                                                                membership_is_active=True):
                        count = 0
                        for question in top_questions:
                            if QuestionResponses.objects.filter(response_question=question,
                                                                response_evaluated=level.membership_level.level_head):
                                count += 1
                        if top_questions.count() == 0:
                            status = 0
                        else:
                            status = count / top_questions.count()

                        if status <= 0:
                            s_tl_not_participated += 1
                        elif 0 < status < 1:
                            s_tl_partially_participated += 1
                        else:
                            s_tl_participated += 1
                            
                        s_tl.append([member_staff, level.membership_level.level_head, status])
                
                
        context['s_tl'] = s_tl
        context['tl_s'] = tl_s
        context['s_tl_participated'] = s_tl_participated
        context['tl_s_participated'] = tl_s_participated

        context['s_tl_partially_participated'] = s_tl_partially_participated
        context['tl_s_partially_participated'] = tl_s_partially_participated

        context['s_tl_not_participated'] = s_tl_not_participated
        context['tl_s_not_participated'] = tl_s_not_participated
        return context


@method_decorator(login_required, name='dispatch')
class ReportCheckIn(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(ReportCheckIn, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        staff = context['staff']

        cat_75 = 0
        cat_50 = 0
        cat_25 = 0
        cat_0 = 0

        my_members_score = []
        levels_down_score = []

        if staff.staff_superuser:
            for member_staff in Staff.objects.filter(staff_company=context['company']):
                cis = CheckIn.objects.filter(check_in_Staff=member_staff, check_in_pms=context['pms'])
                apr = cis.filter(check_in_month="April").count()
                may = cis.filter(check_in_month="May").count()
                jun = cis.filter(check_in_month="June").count()
                jul = cis.filter(check_in_month="July").count()
                aug = cis.filter(check_in_month="August").count()
                sep = cis.filter(check_in_month="September").count()
                oct = cis.filter(check_in_month="October").count()
                nov = cis.filter(check_in_month="November").count()
                dec = cis.filter(check_in_month="December").count()
                jan = cis.filter(check_in_month="January").count()
                feb = cis.filter(check_in_month="February").count()
                mar = cis.filter(check_in_month="March").count()

                months = [apr, may, jun, jul, aug, sep, oct, nov, dec, jan, feb, mar]
                ci_score = calculate_overall_check_in_score(member_staff, context['pms'])

                if ci_score >= 75:
                    cat_75 += 1
                elif 50 <= ci_score < 75:
                    cat_50 += 1
                elif 25 <= ci_score < 50:
                    cat_25 += 1
                else:
                    cat_0 += 1

                my_members_score.append([member_staff, months, ci_score])
        else:
            for level in Level.objects.filter(level_head=staff):
                for members in LevelMembership.objects.filter(membership_level=level):
                    cis = CheckIn.objects.filter(check_in_Staff=members.membership_staff, check_in_pms=context['pms'])
                    apr = cis.filter(check_in_month="April").count()
                    may = cis.filter(check_in_month="May").count()
                    jun = cis.filter(check_in_month="June").count()
                    jul = cis.filter(check_in_month="July").count()
                    aug = cis.filter(check_in_month="August").count()
                    sep = cis.filter(check_in_month="September").count()
                    oct = cis.filter(check_in_month="October").count()
                    nov = cis.filter(check_in_month="November").count()
                    dec = cis.filter(check_in_month="December").count()
                    jan = cis.filter(check_in_month="January").count()
                    feb = cis.filter(check_in_month="February").count()
                    mar = cis.filter(check_in_month="March").count()

                    months = [apr, may, jun, jul, aug, sep, oct, nov, dec, jan, feb, mar]
                    ci_score = calculate_overall_check_in_score(members.membership_staff, context['pms'])

                    if ci_score >= 75:
                        cat_75 += 1
                    elif 50 <= ci_score < 75:
                        cat_50 += 1
                    elif 25 <= ci_score < 50:
                        cat_25 += 1
                    else:
                        cat_0 += 1

                    my_members_score.append([members.membership_staff, months, ci_score])

            if check_staff_is_level_head(self.kwargs['company_id'], staff):
                levels_down = []
                for levels in Level.objects.filter(level_head=staff):
                    all_levels_down(levels, levels_down)

                for level in levels_down:
                    for members in LevelMembership.objects.filter(membership_level=level):
                        cis = CheckIn.objects.filter(check_in_Staff=members.membership_staff,
                                                     check_in_pms=context['pms'])
                        apr = cis.filter(check_in_month="April").count()
                        may = cis.filter(check_in_month="May").count()
                        jun = cis.filter(check_in_month="June").count()
                        jul = cis.filter(check_in_month="July").count()
                        aug = cis.filter(check_in_month="August").count()
                        sep = cis.filter(check_in_month="September").count()
                        oct = cis.filter(check_in_month="October").count()
                        nov = cis.filter(check_in_month="November").count()
                        dec = cis.filter(check_in_month="December").count()
                        jan = cis.filter(check_in_month="January").count()
                        feb = cis.filter(check_in_month="February").count()
                        mar = cis.filter(check_in_month="March").count()

                        months = [apr, may, jun, jul, aug, sep, oct, nov, dec, jan, feb, mar]
                        ci_score = calculate_overall_check_in_score(members.membership_staff, context['pms'])

                        if ci_score >= 75:
                            cat_75 += 1
                        elif 50 <= ci_score < 75:
                            cat_50 += 1
                        elif 25 <= ci_score < 50:
                            cat_25 += 1
                        else:
                            cat_0 += 1

                        levels_down_score.append([members.membership_staff, months, ci_score])

        context['cat_75'] = cat_75
        context['cat_50'] = cat_50
        context['cat_25'] = cat_25
        context['cat_0'] = cat_0
        context['my_members_score'] = my_members_score
        context['levels_down_score'] = levels_down_score
        return context


@method_decorator(login_required, name='dispatch')
class Profile(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(Profile, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        staff = context['staff']
        context['accounts'] = Staff.objects.filter(staff_person=staff.staff_person)
        context['memberships'] = LevelMembership.objects.filter(membership_staff=staff)
        members = []
        heading = Level.objects.filter(level_head=context['staff'])
        for level in heading:
            members.append(LevelMembership.objects.filter(membership_level=level))

        context['heading_levels'] = heading
        context['heading_members'] = members
        return context


@method_decorator(login_required, name='dispatch')
class HelpList(TemplateView):

    def get_context_data(self, **kwargs):
        context = super(HelpList, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        context['help_list'] = Help.objects.filter(help_staff=context['staff'])
        return context


@method_decorator(login_required, name='dispatch')
class HelpCreate(CreateView):
    form_class = HelpForm

    def get_context_data(self, **kwargs):
        context = super(HelpCreate, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        context['help_list'] = Help.objects.filter(help_staff=context['staff'])
        return context

    def get_initial(self):
        initial = super(HelpCreate, self).get_initial()
        initial['help_log_date'] = datetime.datetime.now()
        initial['help_status'] = "Open"
        initial["help_staff"] = get_staff_account(get_company(self.kwargs['company_id']), self.request.user)

        return initial

    def get_success_url(self):
        return '{}'.format(reverse('Site:Help_List', kwargs={'company_id': self.kwargs['company_id']}))


@method_decorator(login_required, name='dispatch')
class HelpEdit(UpdateView):
    model = Help
    form_class = HelpForm

    def get_context_data(self, **kwargs):
        context = super(HelpEdit, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        context['help_list'] = Help.objects.filter(help_staff=context['staff'])
        return context

    def get_success_url(self):
        return '{}'.format(reverse('Site:Help_List', kwargs={'company_id': self.kwargs['company_id']}))


@method_decorator(login_required, name='dispatch')
class HelpDelete(DeleteView):

    def get_context_data(self, **kwargs):
        context = super(HelpDelete, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        context['help_list'] = Help.objects.filter(help_staff=context['staff'])
        return context


@method_decorator(login_required, name='dispatch')
class Communication(TemplateView):
    def get_context_data(self, **kwargs):
        context = super(Communication, self).get_context_data()
        global_context(self.kwargs['company_id'], self.request.user, context)
        context['notification_list'] = Notification.objects.all()

        return context


@login_required
def communicate(request, company_id):
    if request.method == "POST":

        def get_value(x):
            if str(x).strip():
                return str(x).strip()
            else:
                return None

        title = get_value(request.POST.get("communication_title"))
        message = get_value(request.POST.get("communication_message"))

        for staff in Staff.objects.filter(staff_company=get_object_or_404(Company, company_id=company_id)):

            notification_log("Notification", "None", staff.staff_person.get_full_name(), staff.staff_person.email,
                             title, message)

    return HttpResponseRedirect(reverse('Site:Communication', kwargs={'company_id': company_id}))
